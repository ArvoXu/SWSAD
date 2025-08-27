import sys
import os
from datetime import datetime, timedelta
from flask import Flask, jsonify, send_from_directory, request, redirect, render_template, session, flash
from flask_cors import CORS
import schedule
import time
import threading
from dotenv import load_dotenv
from sqlalchemy.orm import Session
from sqlalchemy.exc import OperationalError
from sqlalchemy import or_, and_
import subprocess
import logging
import traceback
from dateutil.parser import parse as parse_date
import pandas as pd
import shutil
import json
import pytz
from openpyxl import load_workbook, Workbook
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
import smtplib
from email.message import EmailMessage


# --- Custom Imports ---
from database import init_db, get_db, Inventory, Store, Transaction, UpdateLog, Warehouse, User, NotificationSent
from scraper import run_scraper as run_inventory_scraper_function, parse_inventory_from_text, save_to_database, save_to_json
from salesscraper import run_sales_scraper
from warehousescraper import run_warehouse_scraper

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Pre-startup Configuration ---
# Get the directory where the script is located
# This is crucial for making file paths work correctly in any environment
script_dir = os.path.dirname(os.path.abspath(__file__))

# Load environment variables from .env file located in the same directory
# This needs to happen before any other part of the app uses them
dotenv_path = os.path.join(script_dir, '.env')
if os.path.exists(dotenv_path):
    logging.info(f"Loading environment variables from: {dotenv_path}")
    load_dotenv(dotenv_path=dotenv_path)
else:
    logging.info(f".env file not found at {dotenv_path}. Relying on system-set environment variables.")


# --- Flask App Setup ---
app = Flask(__name__, 
            static_folder=script_dir, 
            static_url_path='',
            template_folder=os.path.join(script_dir, 'templates')) # Point to the templates folder
CORS(app) # 允許所有來源的跨域請求，方便本地開發

# --- Secret Key for Session Management ---
# It's crucial this is set and kept secret in production.
# We'll load it from an environment variable.
app.secret_key = os.getenv("FLASK_SECRET_KEY", "dev-secret-key-for-local-testing")

# --- Global State for Scraper ---
# This dictionary will hold the state of our scraper job.
# Using a lock to ensure thread-safe updates.
scraper_state = {
    "status": "idle", # Can be 'idle', 'running', 'success', 'error'
    "last_run_output": ""
}
state_lock = threading.Lock()

# --- Global State for Sales Scraper ---
sales_scraper_state = {
    "status": "idle", # Can be 'idle', 'running', 'success', 'error'
    "last_run_output": ""
}
sales_state_lock = threading.Lock()


# --- Global Variables & Constants ---
# Use the script's directory to build absolute paths
# DB_PATH = os.path.join(script_dir, 'inventory.db') # This line is no longer needed
# SCRAPER_SCRIPT_PATH = os.path.join(script_dir, 'scraper.py') # This line is no longer needed
# PYTHON_EXECUTABLE = sys.executable # This line is no longer needed

# --- Database Initialization ---
# Ensure the database and table exist before the app starts serving requests.
# This is crucial for the first run on a new server deployment.
logging.info(f"Initializing database at: {script_dir}")
init_db() # This command creates the tables in our PostgreSQL or SQLite database if they don't exist.
logging.info("Database initialization complete.")


def run_inventory_scraper_background():
    """
    This function runs the scraper in a background thread and updates the global state.
    It now uses the new database functions.
    It's also thread-safe, checking if another job is already running.
    """
    global scraper_state
    
    with state_lock:
        if scraper_state['status'] == 'running':
            logging.warning(f"[{datetime.now()}] Scraper start requested, but a job is already in progress. Aborting.")
            return
        
        # Update state to 'running'
        scraper_state['status'] = 'running'
        scraper_state['last_run_output'] = ''
        logging.info(f"[{datetime.now()}] Scraper status set to 'running'. Starting job.")

    try:
        # We need to call the actual scraper logic here.
        # Since scraper.py's main execution block is complex,
        # we directly call the necessary functions.
        
        # 1. Execute the web scraper to get raw text
        raw_inventory_text = run_inventory_scraper_function(headless=True)
        
        if raw_inventory_text:
            # 2. Parse the raw text into structured data
            structured_data = parse_inventory_from_text(raw_inventory_text)
            
            # --- Define output paths ---
            output_json_path = os.path.join(script_dir, 'structured_inventory.json')
            
            # 3. Save to JSON and Database
            # Convert datetime objects for JSON serialization
            json_serializable_data = []
            for item in structured_data:
                item_copy = item.copy()
                if 'process_time' in item_copy and hasattr(item_copy['process_time'], 'isoformat'):
                    item_copy['process_time'] = item_copy['process_time'].isoformat()
                json_serializable_data.append(item_copy)
            
            save_to_json(json_serializable_data, output_json_path)
            
            # The data is already parsed with datetime objects, so we can save it directly
            items_saved_count = len(structured_data)
            save_to_database(structured_data)

            # After saving inventory to database, immediately run low-inventory notifications
            # Run notifications in a non-blocking daemon thread to avoid delaying the scraper
            try:
                def _notify_runner():
                    try:
                        res = _notify_low_inventory_internal()
                        logging.info(f"Low-inventory notification run after scraper: {res}")
                    except Exception as e:
                        logging.error(f"Error running notifications in background thread: {e}", exc_info=True)

                t = threading.Thread(target=_notify_runner, daemon=True)
                t.start()
                logging.info("Started background notification thread after scraper run.")
            except Exception as e:
                logging.error(f"Failed to start notification thread: {e}", exc_info=True)

            output = f"Scraper finished successfully. Processed {items_saved_count} items. Notifications run." 
            status = "success"
        else:
            output = "Scraper ran but returned no data."
            status = "error"
            
    except Exception as e:
        output = f"An error occurred in background scraper: {str(e)}"
        status = "error"
        logging.error(output, exc_info=True)
    
    # Update state with the final result
    with state_lock:
        scraper_state['status'] = status
        scraper_state['last_run_output'] = output
        
    # Log the update to the database with a retry mechanism
    log_db_update(scraper_type='inventory', status=status, details=output)


def run_sales_scraper_background():
    """
    Runs the sales scraper in a background thread, processes the downloaded file,
    and updates the database with a retry mechanism.
    """
    global sales_scraper_state
    
    with sales_state_lock:
        if sales_scraper_state['status'] == 'running':
            logging.warning(f"[{datetime.now()}] Sales scraper start requested, but a job is already in progress. Aborting.")
            return
        sales_scraper_state['status'] = 'running'
        sales_scraper_state['last_run_output'] = ''
        logging.info(f"[{datetime.now()}] Sales scraper status set to 'running'. Starting job.")

    downloaded_file_path = None
    try:
        # 1. Run the scraper to download the file
        # We explicitly set headless=True to ensure it runs without a GUI on the server.
        downloaded_file_path = run_sales_scraper(headless=True)
        
        # 2. Process the downloaded Excel file
        if downloaded_file_path:
            df = pd.read_excel(downloaded_file_path)
            df.rename(columns={
                'Shop name': 'shopName',
                'Product': 'product',
                'Trasaction Date': 'date',
                'Total Transaction Amount': 'amount',
                'Pay type': 'payType'
            }, inplace=True)
            transactions_data = df.to_dict('records')
            
            # 3. Add transactions to the DB with retry logic
            max_retries = 3
            retry_delay_seconds = 5
            for attempt in range(max_retries):
                db: Session = next(get_db())
                try:
                    # This block is a simplified version of the logic in add_transactions endpoint
                    db.query(Transaction).delete()
                    
                    all_stores = db.query(Store).all()
                    store_name_map = {}
                    for s in all_stores:
                        name_parts = s.store_key.rsplit('-', 1)
                        name = name_parts[0]
                        if name not in store_name_map:
                            store_name_map[name] = s

                    new_transactions = []
                    for item in transactions_data:
                        shop_name_raw = item.get('shopName')
                        if not shop_name_raw or pd.isna(item.get('date')):
                            continue
                        
                        shop_name = str(shop_name_raw).strip()
                        store = store_name_map.get(shop_name)
                        
                        if not store:
                            new_store_key = f"{shop_name}-provisional_sales"
                            store = db.query(Store).filter(Store.store_key == new_store_key).first()
                            if not store:
                                store = Store(store_key=new_store_key)
                                db.add(store)
                                db.flush()
                            store_name_map[shop_name] = store

                        new_transactions.append(Transaction(
                            store_key=store.store_key,
                            transaction_time=pd.to_datetime(item.get('date')),
                            amount=int(float(item.get('amount', 0))),
                            product_name=str(item.get('product')),
                            payment_type=str(item.get('payType'))
                        ))

                    if new_transactions:
                        db.bulk_save_objects(new_transactions)
                    
                    db.commit()
                    processed_count = len(new_transactions)
                    output = f"Sales scraper finished successfully. Processed {processed_count} transactions."
                    status = "success"
                    logging.info(f"Database operation successful on attempt {attempt + 1}.")
                    db.close()
                    break # Exit retry loop on success
                except OperationalError as e:
                    db.rollback()
                    db.close()
                    logging.error(f"Sales DB error (Attempt {attempt + 1}/{max_retries}): {e}")
                    if attempt + 1 >= max_retries:
                        output = f"Sales scraper failed after {max_retries} attempts: {e}"
                        status = "error"
                        raise
                    logging.info(f"Retrying in {retry_delay_seconds} seconds...")
                    time.sleep(retry_delay_seconds)
                finally:
                    # Ensure db is closed if it's still open
                    if 'db' in locals() and db.is_active:
                         db.close()
        else:
            output = "Sales scraper ran but did not return a file path."
            status = "error"
            
    except Exception as e:
        output = f"An error occurred in background sales scraper: {str(e)}"
        status = "error"
        logging.error(output, exc_info=True)
    finally:
        # 4. Clean up downloaded file and temp directory
        if downloaded_file_path:
            download_dir = os.path.dirname(downloaded_file_path)
            try:
                shutil.rmtree(download_dir)
                logging.info(f"Successfully cleaned up temporary directory: {download_dir}")
            except OSError as e:
                logging.error(f"Error removing directory {download_dir}: {e.strerror}")
                
        with sales_state_lock:
            sales_scraper_state['status'] = status
            sales_scraper_state['last_run_output'] = output
            
        # Log the update to the database with a retry mechanism
        log_db_update(scraper_type='sales', status=status, details=output)


def log_db_update(scraper_type, status, details):
    """Logs an update record to the database with a retry mechanism."""
    max_retries = 3
    retry_delay = 5
    for attempt in range(max_retries):
        db_log: Session = next(get_db())
        try:
            log_entry = UpdateLog(scraper_type=scraper_type, status=status, details=details)
            db_log.add(log_entry)
            db_log.commit()
            logging.info(f"Successfully logged '{status}' for '{scraper_type}' scraper.")
            return
        except OperationalError as e:
            db_log.rollback()
            logging.error(f"Failed to write to update_logs (Attempt {attempt + 1}/{max_retries}): {e}")
            if attempt + 1 >= max_retries:
                logging.error(f"Giving up on logging after {max_retries} attempts.")
            else:
                logging.info(f"Retrying log write in {retry_delay} seconds...")
                time.sleep(retry_delay)
        except Exception as e:
            db_log.rollback()
            logging.error(f"An unexpected error occurred while writing to update_logs: {e}", exc_info=True)
            return # Don't retry on unexpected errors
        finally:
            db_log.close()


# --- Warehouse Scraper Background Function ---
def run_warehouse_scraper_background():
    """
    在背景執行倉庫爬蟲，處理下載的檔案並更新資料庫。
    """
    global sales_scraper_state
    
    with sales_state_lock:
        if sales_scraper_state['status'] == 'running':
            logging.warning(f"[{datetime.now()}] Warehouse scraper start requested, but a job is already in progress. Aborting.")
            return
        sales_scraper_state['status'] = 'running'
        sales_scraper_state['last_run_output'] = ''
        logging.info(f"[{datetime.now()}] Warehouse scraper status set to 'running'. Starting job.")

    downloaded_file_path = None
    try:
        # 1. 執行爬蟲下載檔案
        downloaded_file_path = run_warehouse_scraper(headless=True)
        
        # 2. 處理下載的 Excel 檔案
        if downloaded_file_path:
            df = pd.read_excel(downloaded_file_path)
            
            # 確保必要的欄位存在
            required_columns = ['Warehouse name', 'Product name', 'Remain quantity']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                raise ValueError(f"Excel 檔案缺少必要欄位: {', '.join(missing_columns)}")
            
            # 更新資料庫
            max_retries = 3
            retry_delay_seconds = 5
            
            for attempt in range(max_retries):
                db: Session = next(get_db())
                try:
                    # 記錄當前時間作為更新時間
                    update_time = datetime.now()
                    
                    # 將資料轉換為資料庫記錄
                    warehouse_records = []
                    for _, row in df.iterrows():
                        warehouse_records.append(Warehouse(
                            warehouse_name=row['Warehouse name'],
                            product_name=row['Product name'],
                            quantity=int(row['Remain quantity']),
                            updated_at=update_time
                        ))
                    
                    # 刪除舊的倉庫資料
                    db.query(Warehouse).delete()
                    
                    # 新增新的倉庫資料
                    db.bulk_save_objects(warehouse_records)
                    db.commit()
                    
                    output = f"成功更新倉庫資料。處理了 {len(warehouse_records)} 筆記錄。"
                    status = "success"
                    break
                    
                except OperationalError as e:
                    db.rollback()
                    if attempt + 1 >= max_retries:
                        raise
                    logging.error(f"資料庫操作失敗 (嘗試 {attempt + 1}/{max_retries}): {e}")
                    time.sleep(retry_delay_seconds)
                finally:
                    db.close()
        else:
            output = "倉庫爬蟲執行完成但未返回檔案路徑。"
            status = "error"
            
    except Exception as e:
        output = f"倉庫爬蟲過程中發生錯誤: {str(e)}"
        status = "error"
        logging.error(output, exc_info=True)
    finally:
        # 清理下載的檔案和暫存目錄
        if downloaded_file_path:
            download_dir = os.path.dirname(downloaded_file_path)
            try:
                shutil.rmtree(download_dir)
                logging.info(f"成功清理暫存目錄: {download_dir}")
            except OSError as e:
                logging.error(f"移除目錄時發生錯誤 {download_dir}: {e.strerror}")
                
        with sales_state_lock:
            sales_scraper_state['status'] = status
            sales_scraper_state['last_run_output'] = output
            
        # 記錄更新到資料庫
        log_db_update(scraper_type='warehouse', status=status, details=output)

# --- API Endpoints ---
@app.route('/run-scraper', methods=['POST'])
def trigger_scraper():
    """
    Triggers the scraper script to run in a background thread.
    Returns immediately so the client doesn't time out.
    """
    with state_lock:
        if scraper_state['status'] == 'running':
            return jsonify({'success': False, 'message': 'Scraper is already running.'}), 409

    logging.info(f"[{datetime.now()}] Received request to run scraper in background.")
    # Run the scraper in a separate thread
    thread = threading.Thread(target=run_inventory_scraper_background)
    thread.start()
    
    return jsonify({'success': True, 'message': 'Scraper job started in the background.'}), 202

@app.route('/upload-inventory-file', methods=['POST'])
def upload_inventory_file():
    """
    接收格式化好的庫存數據文件（JSON格式）並直接保存到數據庫
    避免在伺服器上運行爬蟲腳本，減少CPU負載
    """
    try:
        # 檢查是否有文件上傳
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '沒有選擇文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '沒有選擇文件'}), 400
        
        # 檢查文件類型
        if not file.filename.endswith('.json'):
            return jsonify({'success': False, 'message': '只支持 JSON 格式的文件'}), 400
        
        # 讀取文件內容
        file_content = file.read()
        try:
            inventory_data = json.loads(file_content.decode('utf-8'))
        except json.JSONDecodeError as e:
            return jsonify({'success': False, 'message': f'JSON 格式錯誤: {str(e)}'}), 400
        
        # 驗證數據格式
        if not isinstance(inventory_data, list):
            return jsonify({'success': False, 'message': '數據格式錯誤：應該是列表格式'}), 400
        
        # 驗證每個項目是否包含必要字段
        required_fields = ['store', 'machine_id', 'product_name', 'quantity']
        for i, item in enumerate(inventory_data):
            if not isinstance(item, dict):
                return jsonify({'success': False, 'message': f'第 {i+1} 項數據格式錯誤：應該是字典格式'}), 400
            
            missing_fields = [field for field in required_fields if field not in item]
            if missing_fields:
                return jsonify({'success': False, 'message': f'第 {i+1} 項缺少必要字段: {", ".join(missing_fields)}'}), 400
            
            # 驗證數量字段
            if not isinstance(item['quantity'], int) or item['quantity'] < 0:
                return jsonify({'success': False, 'message': f'第 {i+1} 項數量字段錯誤：應該是正整數'}), 400
        
        # 處理日期時間字段
        for item in inventory_data:
            if 'process_time' in item:
                if isinstance(item['process_time'], str):
                    try:
                        item['process_time'] = parse_date(item['process_time'])
                    except Exception as e:
                        return jsonify({'success': False, 'message': f'日期時間格式錯誤: {str(e)}'}), 400
            else:
                # 如果沒有 process_time，使用當前時間
                item['process_time'] = datetime.now(pytz.timezone('Asia/Taipei'))
        
        # 保存到數據庫
        max_retries = 3
        retry_delay_seconds = 5
        
        for attempt in range(max_retries):
            db: Session = next(get_db())
            try:
                # 清空現有庫存數據
                num_deleted = db.query(Inventory).delete()
                logging.info(f"Cleared {num_deleted} old records from the inventory table.")
                
                # 創建新的庫存對象
                inventory_objects = [Inventory(**item) for item in inventory_data]
                db.bulk_save_objects(inventory_objects)
                
                db.commit()
                items_saved_count = len(inventory_objects)
                logging.info(f"Successfully saved {items_saved_count} new records to the database via file upload.")
                
                # 記錄更新日誌
                log_db_update(
                    scraper_type='inventory_upload', 
                    status='success', 
                    details=f'File upload successful. Processed {items_saved_count} items from {file.filename}'
                )
                
                db.close()
                return jsonify({
                    'success': True, 
                    'message': f'成功上傳並處理 {items_saved_count} 項庫存數據',
                    'items_processed': items_saved_count,
                    'filename': file.filename
                })
                
            except OperationalError as e:
                db.rollback()
                db.close()
                logging.error(f"Database error on attempt {attempt + 1}/{max_retries}: {e}")
                if attempt + 1 >= max_retries:
                    log_db_update(
                        scraper_type='inventory_upload', 
                        status='error', 
                        details=f'Database error after {max_retries} attempts: {str(e)}'
                    )
                    return jsonify({'success': False, 'message': f'數據庫錯誤，已重試 {max_retries} 次: {str(e)}'}), 500
                logging.info(f"Retrying in {retry_delay_seconds} seconds...")
                time.sleep(retry_delay_seconds)
                
            except Exception as e:
                db.rollback()
                db.close()
                logging.error(f"Unexpected error during file upload: {e}", exc_info=True)
                log_db_update(
                    scraper_type='inventory_upload', 
                    status='error', 
                    details=f'Unexpected error: {str(e)}'
                )
                return jsonify({'success': False, 'message': f'處理文件時發生錯誤: {str(e)}'}), 500
            finally:
                if 'db' in locals() and db.is_active:
                    db.close()
                    
    except Exception as e:
        logging.error(f"Error in file upload endpoint: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'文件上傳失敗: {str(e)}'}), 500

@app.route('/scraper-status', methods=['GET'])
def get_scraper_status():
    """
    Returns the current status of the scraper job.
    """
    with state_lock:
        return jsonify(scraper_state)

@app.route('/run-sales-scraper', methods=['POST'])
def trigger_sales_scraper():
    """
    Triggers the sales scraper to run in a background thread.
    """
    with sales_state_lock:
        if sales_scraper_state['status'] == 'running':
            return jsonify({'success': False, 'message': 'Sales scraper is already running.'}), 409

    logging.info(f"[{datetime.now()}] Received request to run sales scraper in background.")
    thread = threading.Thread(target=run_sales_scraper_background)
    thread.start()
    
    return jsonify({'success': True, 'message': 'Sales scraper job started in the background.'}), 202

@app.route('/sales-scraper-status', methods=['GET'])
def get_sales_scraper_status():
    """
    Returns the current status of the sales scraper job.
    """
    with sales_state_lock:
        return jsonify(sales_scraper_state)

def to_camel_case(snake_str):
    """
    Converts a snake_case string to camelCase.
    Example: 'product_name' -> 'productName'
    """
    components = snake_str.split('_')
    # Capitalize the first letter of each component except the first one
    # and join them together.
    return components[0] + ''.join(x.title() for x in components[1:])

@app.route('/get-data', methods=['GET'])
def get_data():
    """
    Retrieves all inventory and custom store data, merges them,
    and returns them as a single JSON response with camelCase keys.
    """
    db: Session = next(get_db())
    try:
        # If a normal user is logged in (session['user_id']), restrict results to their assigned stores
        user_id = session.get('user_id')
        if user_id:
            user = db.query(User).filter(User.id == int(user_id)).first()
            allowed_store_keys = set()
            if user:
                # user's stores contain store_key strings
                allowed_store_keys = set(s.store_key for s in user.stores)
            # Query inventory with flexible matching: allow matching by
            # 1) exact store + machine (if applicable),
            # 2) Inventory.store equals the prefix part of store_key,
            # 3) Inventory.store LIKE prefix% fallback.
            inventory_items = []
            filters = []
            for sk in allowed_store_keys:
                # if store_key looks like 'store-machine', try exact match first
                if '-' in sk:
                    left, right = sk.rsplit('-', 1)
                    # exact store+machine
                    filters.append(and_(Inventory.store == left, Inventory.machine_id == right))
                    # any machine for that store (case-insensitive)
                    filters.append(Inventory.store.ilike(left))
                    # fallback patterns: startswith and contains (case-insensitive)
                    filters.append(Inventory.store.ilike(f"{left}%"))
                    filters.append(Inventory.store.ilike(f"%{left}%"))
                else:
                    filters.append(Inventory.store.ilike(sk))
                    filters.append(Inventory.store.ilike(f"{sk}%"))
                    filters.append(Inventory.store.ilike(f"%{sk}%"))
            if filters:
                # combine with OR and query
                inventory_items = db.query(Inventory).filter(or_(*filters)).all()
        else:
            inventory_items = db.query(Inventory).all()
        
        # 2. Fetch all custom store data
        stores = db.query(Store).all()
        # Create a dictionary for quick lookups: {'store_key': {address: '...', 'note': '...'}}
        store_info_map = {store.store_key: store.to_dict() for store in stores}

        # 3. Merge the data
        merged_data = []
        for item in inventory_items:
            item_dict = item.to_dict()
            store_key = f"{item.store}-{item.machine_id}"
            
            # Get custom data for this store, if it exists
            custom_store_data = store_info_map.get(store_key, {})
            
            # Merge inventory data with custom store data
            full_item_data = {**item_dict, **custom_store_data}
            
            # 4. Convert all keys to camelCase for the frontend
            camel_case_data = {to_camel_case(key): value for key, value in full_item_data.items()}
            
            # Ensure process_time is in ISO format string
            if 'processTime' in camel_case_data and hasattr(camel_case_data['processTime'], 'isoformat'):
                camel_case_data['processTime'] = camel_case_data['processTime'].isoformat()

            merged_data.append(camel_case_data)
        
        return jsonify({"success": True, "data": merged_data})
        
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        db.close()


def is_admin():
    """Helper: current session indicates admin if session['logged_in']==True (existing behavior)."""
    return bool(session.get('logged_in'))


@app.route('/api/stores-list', methods=['GET'])
def api_stores_list():
    """Return list of stores for admin forms.

    By default this endpoint returns distinct `Inventory.store` values (so admins
    can assign users based on the inventory naming). If callers want the older
    `Store.store_key` list, pass ?source=stores.
    """
    source = (request.args.get('source') or 'inventory').lower()
    db: Session = next(get_db())
    try:
        if source == 'stores':
            stores = db.query(Store).all()
            result = [s.store_key for s in stores]
        else:
            # default: distinct Inventory.store values
            rows = db.query(Inventory.store).distinct().order_by(Inventory.store).all()
            # rows is list of 1-tuples like [('台北天文館 左邊',), ...]
            result = [r[0] for r in rows if r and r[0]]
        return jsonify({'success': True, 'stores': result})
    except Exception as e:
        logging.error(f"Error getting stores list: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        db.close()


@app.route('/api/users', methods=['POST'])
def api_create_user():
    """Admin-only: create a user and assign stores. Request JSON: {username, password, displayName, stores: []} """
    if not is_admin():
        return jsonify({'success': False, 'message': '未授權'}), 403

    data = request.get_json() or {}
    username = data.get('username')
    password = data.get('password')
    display_name = data.get('displayName') or ''
    stores = data.get('stores') or []

    if not username or not password:
        return jsonify({'success': False, 'message': 'username 和 password 為必填'}), 400

    db: Session = next(get_db())
    try:
        existing = db.query(User).filter(User.username == username).first()
        if existing:
            return jsonify({'success': False, 'message': 'username 已存在'}), 400

        hashed = generate_password_hash(password)
        user = User(username=username, password_hash=hashed, display_name=display_name)
        # assign stores if exist. Accept either a Store.store_key (legacy) or
        # an Inventory.store name. We try to map the provided value to an
        # existing Store by exact match, prefix match, or by creating a
        # provisional store_key if nothing matches.
        for sk in stores:
            store = db.query(Store).filter(Store.store_key == sk).first()
            if not store:
                # try prefix/pattern match (e.g., inventory name matches the
                # left part of store_key like '台北天文館 左邊' -> '台北天文館 左邊-...')
                store = db.query(Store).filter(Store.store_key.like(f"{sk}%")).first()
            if not store:
                # fallback: create a provisional store entry using a common suffix
                new_key = f"{sk}-provisional_sales"
                store = db.query(Store).filter(Store.store_key == new_key).first()
                if not store:
                    store = Store(store_key=new_key)
                    db.add(store)
                    db.flush()
            if store:
                user.stores.append(store)

        db.add(user)
        db.commit()
        db.refresh(user)
        return jsonify({'success': True, 'user': user.to_dict()})
    except Exception as e:
        db.rollback()
        logging.error(f"Error creating user: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        db.close()


@app.route('/api/user-login', methods=['POST'])
def api_user_login():
    """User login for presentation page. JSON: {username, password} sets session['user_id'] on success."""
    data = request.get_json() or {}
    username = data.get('username')
    password = data.get('password')
    if not username or not password:
        return jsonify({'success': False, 'message': 'username/password required'}), 400

    db: Session = next(get_db())
    try:
        user = db.query(User).filter(User.username == username).first()
        if not user or not check_password_hash(user.password_hash, password):
            return jsonify({'success': False, 'message': '帳號或密碼錯誤'}), 401
        # set session user id; keep admin separate
        session['user_id'] = user.id
        return jsonify({'success': True, 'user': user.to_dict()})
    except Exception as e:
        logging.error(f"Error during user login: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        db.close()


@app.route('/api/user-logout', methods=['POST'])
def api_user_logout():
    session.pop('user_id', None)
    return jsonify({'success': True})


@app.route('/api/user-info', methods=['GET'])
def api_user_info():
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'logged_in': False}), 200
    db: Session = next(get_db())
    try:
        user = db.query(User).filter(User.id == int(user_id)).first()
        if not user:
            return jsonify({'logged_in': False}), 200
        return jsonify({'logged_in': True, 'user': user.to_dict()})
    except Exception as e:
        logging.error(f"Error fetching user info: {e}")
        return jsonify({'logged_in': False}), 500
    finally:
        db.close()


@app.route('/presentation-login', methods=['GET'])
def presentation_login_page():
    """Serve a standalone presentation login page for franchisee users.

    The page posts to /api/user-login via fetch and redirects back to
    /presentation.html on success. A `next` query parameter is accepted.
    """
    return send_from_directory(script_dir, 'presentation_login.html')


@app.route('/presentation-logout', methods=['POST'])
def presentation_logout():
    """Optional helper to log out a franchisee and return JSON."""
    session.pop('user_id', None)
    return jsonify({'success': True})


@app.route('/api/user-profile', methods=['GET', 'POST'])
def api_user_profile():
    """Get or update the current logged-in user's profile (display_name, email).

    GET -> {logged_in, user}
    POST -> accepts JSON {displayName, email} and updates DB
    """
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': '未登入'}), 401

    db: Session = next(get_db())
    try:
        user = db.query(User).filter(User.id == int(user_id)).first()
        if not user:
            return jsonify({'success': False, 'message': 'user not found'}), 404

        if request.method == 'GET':
            return jsonify({'success': True, 'user': user.to_dict()})

        # POST -> update
        data = request.get_json() or {}
        display_name = data.get('displayName')
        email = data.get('email')
        low_threshold = data.get('lowInventoryThreshold')
        changed = False
        if display_name is not None:
            user.display_name = display_name
            changed = True
        if email is not None:
            user.email = email
            changed = True
        if low_threshold is not None:
            try:
                user.low_inventory_threshold = int(low_threshold)
                changed = True
            except Exception:
                pass
        if changed:
            db.add(user)
            db.commit()
            db.refresh(user)
        return jsonify({'success': True, 'user': user.to_dict()})
    except Exception as e:
        db.rollback()
        logging.error(f"Error in user-profile: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        db.close()


@app.route('/api/user-change-password', methods=['POST'])
def api_user_change_password():
    """Change the logged-in user's password. JSON {currentPassword, newPassword}.
    Verifies current password and updates stored hash.
    """
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'message': '未登入'}), 401

    data = request.get_json() or {}
    current = data.get('currentPassword') or ''
    new = data.get('newPassword') or ''
    if not current or not new:
        return jsonify({'success': False, 'message': 'currentPassword and newPassword are required'}), 400
    if len(new) < 8:
        return jsonify({'success': False, 'message': '新密碼長度至少 8 字元'}), 400

    db: Session = next(get_db())
    try:
        user = db.query(User).filter(User.id == int(user_id)).first()
        if not user:
            return jsonify({'success': False, 'message': 'user not found'}), 404
        # verify current password
        if not check_password_hash(user.password_hash, current):
            return jsonify({'success': False, 'message': '目前密碼錯誤'}), 403

        # update
        user.password_hash = generate_password_hash(new)
        db.add(user)
        db.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.rollback()
        logging.error(f"Error changing password: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        db.close()


def send_email_if_configured(to_email, subject, body, html_content=False):
    smtp_server = os.getenv('SMTP_SERVER')
    smtp_port = int(os.getenv('SMTP_PORT', '587'))
    smtp_user = os.getenv('SMTP_USER')
    smtp_pass = os.getenv('SMTP_PASS')
    from_email = os.getenv('FROM_EMAIL') or smtp_user
    if not smtp_server or not to_email or not from_email:
        logging.warning(f"Email not sent (missing config): smtp={smtp_server} to={to_email} from={from_email}")
        return {'ok': False, 'error': 'SMTP or from/to address missing'}
    try:
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = from_email
        msg['To'] = to_email
        # Support HTML content when requested. Provide a plain-text fallback.
        if html_content:
            # Simple plain-text fallback; avoid stripping HTML aggressively here.
            fallback = '此郵件包含 HTML 內容，請使用支援 HTML 的郵件客戶端查看。'
            msg.set_content(fallback)
            msg.add_alternative(body, subtype='html')
        else:
            msg.set_content(body)

        # Connect and send via TLS (SendGrid uses TLS on 587)
        with smtplib.SMTP(smtp_server, smtp_port, timeout=10) as s:
            s.ehlo()
            try:
                s.starttls()
                s.ehlo()
            except Exception:
                logging.debug('starttls not supported or failed, continuing without it')
            if smtp_user and smtp_pass:
                try:
                    s.login(smtp_user, smtp_pass)
                except Exception as e:
                    logging.error(f"SMTP login failed for {smtp_user}: {e}")
                    return {'ok': False, 'error': f'SMTP login failed: {e}'}
            s.send_message(msg)
        logging.info(f"Sent email to {to_email} from {from_email} via {smtp_server}:{smtp_port}")
        return {'ok': True}
    except Exception as e:
        logging.error(f"Failed to send email to {to_email}: {e}", exc_info=True)
        return {'ok': False, 'error': str(e)}


@app.route('/api/notify-low-inventory', methods=['POST'])
def api_notify_low_inventory():
    """Manual trigger: check all users and send notification to those below threshold.

    This can be used in a scheduled job; here it's exposed for testing.
    """
    result = _notify_low_inventory_internal()
    if result.get('error'):
        return jsonify({'success': False, 'message': result['error']}), 500
    return jsonify({'success': True, 'notifications': result.get('notifications', [])})
    

def _notify_low_inventory_internal():
    db: Session = next(get_db())
    notifications = []
    # Daily limits and tracking
    DAILY_EMAIL_LIMIT = int(os.getenv('DAILY_EMAIL_LIMIT', '100'))
    emails_sent_today = 0
    # compute today's reset boundary at 16:00 UTC (which corresponds to Taiwan midnight)
    # This means notifications are considered "same day" if sent after the previous 16:00 UTC.
    from datetime import datetime, timedelta
    now_utc = datetime.utcnow()
    # reset happens at 16:00 UTC each UTC day
    reset_hour_utc = 16
    today_start = now_utc.replace(hour=reset_hour_utc, minute=0, second=0, microsecond=0)
    if now_utc.hour < reset_hour_utc:
        # If current UTC time is before 16:00, the 'today' start is the previous day's 16:00
        today_start = today_start - timedelta(days=1)
    try:
        users = db.query(User).all()
        logging.info(f'Checking low-inventory for {len(users)} users')
        for u in users:
            try:
                thresh = int(u.low_inventory_threshold or 0)
            except Exception:
                thresh = 0
            if not u.email or thresh <= 0:
                logging.debug(f"Skipping user {u.username}: no email or threshold={thresh}")
                continue
            # Check each assigned store individually (single-machine total), and notify if any store <= threshold
            low_stores = []
            for s in u.stores:
                # Respect the authoritative Store.is_hidden value from the DB (admin may toggle this)
                try:
                    db_store = db.query(Store).filter(Store.store_key == s.store_key).first()
                except Exception:
                    db_store = None

                if db_store and getattr(db_store, 'is_hidden', False):
                    logging.debug(f"Skipping hidden store {s.store_key} for user {u.username} because is_hidden")
                    continue

                left = s.store_key.split('-')[0]
                qty_rows = db.query(Inventory).filter(Inventory.store.ilike(f"%{left}%")).with_entities(Inventory.quantity).all()
                total_for_store = sum([q[0] or 0 for q in qty_rows])

                # Only log and include stores that are actually below threshold
                if total_for_store <= thresh:
                    logging.info(f"User {u.username} store {s.store_key} totalQuantity={total_for_store} threshold={thresh}")
                    # Prepare display name without provisional suffix
                    display_name = s.store_key
                    if display_name.endswith('-provisional_sales'):
                        display_name = display_name.replace('-provisional_sales', '')
                    low_stores.append({'store_key': s.store_key, 'display': display_name, 'total': total_for_store})

            if low_stores:
                # Filter out stores that have already triggered a notification for this user today
                filtered_low_stores = []
                for ls in low_stores:
                    already = db.query(NotificationSent).filter(
                        NotificationSent.user_id == u.id,
                        NotificationSent.store_key == ls['store_key'],
                        NotificationSent.sent_at >= today_start
                    ).first()
                    if not already:
                        filtered_low_stores.append(ls)

                if not filtered_low_stores:
                    logging.debug(f"User {u.username}: all low stores already notified today; skipping.")
                    continue

                # Check global daily limit
                if emails_sent_today >= DAILY_EMAIL_LIMIT:
                    logging.warning(f"Daily email limit reached ({DAILY_EMAIL_LIMIT}). Skipping notifications.")
                    break

                # We'll send one email per user containing ALL current low_stores (for context),
                # but only record NotificationSent for the newly-notified stores (filtered_low_stores).
                # Compose a single email listing all low stores for this user
                subject = f"【SWSAD】庫存通知 - {len(filtered_low_stores)} 個機台需要您的關注"

                # 使用 HTML 格式
                body_html_lines = [
                    f"<!DOCTYPE html><html><body>",
                    f"<h1 style='text-align:center; font-weight:bold;'>SWSAD</h1>",
                    f"<p>親愛的 {u.display_name or u.username}，</p>",
                    f"<p>這是一則來自SWSAD小幫手的通知：</p>",
                    f"<p>系統發現有幾台機器的庫存已經低於您設定的警戒值囉！為了確保銷售不中斷，建議您盡快安排補貨。</p>",
                    
                    # 開始建立表格
                    f"<table style='width:100%; border-collapse:collapse; text-align:left;'>",
                    f"   <tr style='background-color:#f2f2f2;'>",
                    f"       <th style='padding:8px; border:1px solid #ddd;'>機台名稱</th>",
                    f"       <th style='padding:8px; border:1px solid #ddd;'>目前庫存</th>",
                    f"   </tr>"
                ]

                for ls in low_stores:
                    # 每一行資料
                    body_html_lines.append(f"<tr>")
                    body_html_lines.append(f"    <td style='padding:8px; border:1px solid #ddd;'>{ls.get('display', ls['store_key'])}</td>")
                    body_html_lines.append(f"    <td style='padding:8px; border:1px solid #ddd;'>{ls['total']} 個</td>")
                    body_html_lines.append(f"</tr>")
                    
                body_html_lines.extend([
                    f"</table>", # 表格結束
                    
                    f"<p>點擊下方連結，即可前往網站查看詳細庫存狀況並安排補貨：</p>",
                    f'<p><a href="https://swsad3.onrender.com/presentation">👉 智慧倉儲與銷售分析儀表板-Smart Warehousing and Sales Analysis Dashboard</a></p>',
                    f"<p>此為系統自動通知，請勿回覆。</p>",
                    f"</body></html>"
                ])

                body_html = "\n".join(body_html_lines)
                send_result = send_email_if_configured(u.email, subject, body_html, html_content=True)
                # If sent successfully (or attempted), record sent notifications for each newly-notified store only
                try:
                    for ls in filtered_low_stores:
                        ns = NotificationSent(user_id=u.id, store_key=ls['store_key'])
                        db.add(ns)
                    db.commit()
                except Exception as e:
                    db.rollback()
                    logging.error(f"Failed to record NotificationSent: {e}", exc_info=True)

                emails_sent_today += 1
                notifications.append({'user': u.username, 'email': u.email, 'lowStores': filtered_low_stores, 'allLowStores': low_stores, 'threshold': thresh, 'sent': send_result})
            else:
                logging.debug(f"User {u.username} has no assigned stores below threshold {thresh}")
        return {'notifications': notifications}
    except Exception as e:
        logging.error(f"Error in notify-low-inventory internal: {e}", exc_info=True)
        return {'error': str(e)}
    finally:
        db.close()


@app.route('/api/notify-low-inventory/trigger', methods=['GET'])
def api_notify_low_inventory_trigger():
    """Convenience GET endpoint to trigger low-inventory notifications (for testing via browser)."""
    result = _notify_low_inventory_internal()
    if result.get('error'):
        return jsonify({'success': False, 'message': result['error']}), 500
    return jsonify({'success': True, 'notifications': result.get('notifications', [])})


# Serve presentation page but redirect unauthenticated users to the presentation login
@app.route('/presentation.html', methods=['GET'])
def presentation_page():
    user_id = session.get('user_id')
    if not user_id:
        # Redirect to presentation-login and include next param
        return redirect(f"/presentation-login?next=/presentation.html")
    return send_from_directory(script_dir, 'presentation.html')


@app.route('/presentation', methods=['GET'])
def presentation_short():
    return redirect('/presentation.html')

@app.route('/api/update-logs', methods=['GET'])
def get_update_logs():
    """Returns the last 50 update log entries, newest first."""
    db: Session = next(get_db())
    try:
        logs = db.query(UpdateLog).order_by(UpdateLog.ran_at.desc()).limit(50).all()
        result = [
            {
                "scraperType": log.scraper_type,
                "ranAt": log.ran_at.isoformat(),
                "status": log.status,
                "details": log.details
            }
            for log in logs
        ]
        return jsonify(result)
    except Exception as e:
        logging.error(f"Error getting update logs: {e}", exc_info=True)
        return jsonify({"error": "Could not retrieve update logs"}), 500
    finally:
        db.close()

@app.route('/api/stores/<string:store_key>', methods=['POST'])
def update_store_data(store_key):
    """
    Updates the custom data for a specific store (address, note, sales, hidden status).
    This is the new endpoint for saving user edits.
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "No data provided"}), 400

    db: Session = next(get_db())
    try:
        store = db.query(Store).filter(Store.store_key == store_key).first()
        
        if not store:
            store = Store(store_key=store_key)
            db.add(store)
        
        if 'address' in data:
            store.address = data['address']
        if 'note' in data:
            store.note = data['note']
        if 'manualSales' in data:
            store.manual_sales = data['manualSales']
        if 'isHidden' in data:
            store.is_hidden = data['isHidden']
            
        db.commit()
        db.refresh(store)
        
        return jsonify({"success": True, "data": store.to_dict()})
        
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        db.close()

@app.route('/api/transactions', methods=['POST'])
def add_transactions():
    """
    Receives a list of transactions, clears existing ones, and saves the new ones.
    Includes a retry mechanism for database operations.
    """
    transactions_data = request.get_json()
    if not isinstance(transactions_data, list):
        return jsonify({"success": False, "message": "Invalid data format. Expected a list of transactions."}), 400

    max_retries = 3
    retry_delay_seconds = 5
    for attempt in range(max_retries):
        db: Session = next(get_db())
        try:
            # Clear existing transactions
            db.query(Transaction).delete()
            logging.info("Cleared existing transactions.")
            
            new_transactions = []
            
            all_stores = db.query(Store).all()
            store_name_map = {}
            for s in all_stores:
                name_parts = s.store_key.rsplit('-', 1)
                name = name_parts[0]
                if name not in store_name_map:
                    store_name_map[name] = s

            for item in transactions_data:
                shop_name_raw = item.get('shopName')
                if not shop_name_raw or not item.get('date'):
                    continue
                
                shop_name = shop_name_raw.strip()
                store = store_name_map.get(shop_name)
                
                if not store:
                    logging.info(f"Shop '{shop_name}' not found in DB. Creating a new provisional store.")
                    new_store_key = f"{shop_name}-provisional_sales"
                    
                    existing_provisional = db.query(Store).filter(Store.store_key == new_store_key).first()
                    if existing_provisional:
                        store = existing_provisional
                    else:
                        store = Store(store_key=new_store_key)
                        db.add(store)
                        db.flush()
                    
                    store_name_map[shop_name] = store

                new_transactions.append(Transaction(
                    store_key=store.store_key,
                    transaction_time=parse_date(item.get('date')),
                    amount=int(float(item.get('amount', 0))),
                    product_name=item.get('product'),
                    payment_type=item.get('payType')
                ))

            if new_transactions:
                db.bulk_save_objects(new_transactions)
                logging.info(f"Preparing to save {len(new_transactions)} new transactions.")
            
            db.commit()
            logging.info("Successfully committed transactions.")
            db.close()
            return jsonify({"success": True, "message": f"Successfully added {len(new_transactions)} transactions."})

        except OperationalError as e:
            db.rollback()
            db.close()
            logging.error(f"DB Error on attempt {attempt + 1}: {e}")
            if attempt + 1 >= max_retries:
                logging.error("Max retries reached. Aborting.")
                return jsonify({"success": False, "message": f"Database error after {max_retries} attempts: {e}"}), 500
            logging.info(f"Retrying in {retry_delay_seconds} seconds...")
            time.sleep(retry_delay_seconds)
        except Exception as e:
            db.rollback()
            db.close()
            logging.error(f"Error adding transactions: {e}")
            traceback.print_exc()
            return jsonify({"success": False, "message": str(e)}), 500
        finally:
            if 'db' in locals() and db.is_active:
                db.close()


@app.route('/upload-warehouse-file', methods=['POST'])
def upload_warehouse_file():
    """
    接收倉庫庫存 Excel 文件並保存到數據庫
    只替換同名倉庫的數據，不同倉庫的數據會保留
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': '沒有選擇文件'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': '沒有選擇文件'}), 400
        
        # 檢查文件類型
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': '只支持 Excel 格式的文件 (.xlsx, .xls)'}), 400
        
        # 讀取 Excel 文件
        df = pd.read_excel(file)
        required_columns = ['Warehouse name', 'Product name', 'Remain quantity']
        
        # 驗證必要欄位
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False, 
                'message': f'缺少必要欄位: {", ".join(missing_columns)}'
            }), 400
        
        # 資料處理和驗證
        df['Remain quantity'] = pd.to_numeric(df['Remain quantity'], errors='coerce')
        df = df.dropna(subset=['Warehouse name', 'Product name', 'Remain quantity'])
        
        # 更新資料庫
        max_retries = 3
        retry_delay_seconds = 5
        
        for attempt in range(max_retries):
            db: Session = next(get_db())
            try:
                # 獲取要更新的倉庫列表
                warehouse_names = df['Warehouse name'].unique()
                
                # 只刪除要更新的倉庫的數據
                for warehouse in warehouse_names:
                    db.query(Warehouse).filter(Warehouse.warehouse_name == warehouse).delete()
                
                # 創建新的倉庫記錄
                warehouse_records = []
                for _, row in df.iterrows():
                    warehouse_records.append(Warehouse(
                        warehouse_name=str(row['Warehouse name']),
                        product_name=str(row['Product name']),
                        quantity=int(row['Remain quantity']),
                        updated_at=datetime.now(pytz.timezone('Asia/Taipei'))
                    ))
                
                db.bulk_save_objects(warehouse_records)
                db.commit()
                
                return jsonify({
                    'success': True,
                    'message': f'成功上傳並處理 {len(warehouse_records)} 筆倉庫數據',
                    'records_count': len(warehouse_records)
                })
                
            except OperationalError as e:
                db.rollback()
                logging.error(f"Database error on attempt {attempt + 1}/{max_retries}: {e}")
                if attempt + 1 >= max_retries:
                    return jsonify({'success': False, 'message': f'數據庫錯誤: {str(e)}'}), 500
                time.sleep(retry_delay_seconds)
            except Exception as e:
                db.rollback()
                logging.error(f"Error processing warehouse file: {e}", exc_info=True)
                return jsonify({'success': False, 'message': f'處理文件時發生錯誤: {str(e)}'}), 500
            finally:
                db.close()
                
    except Exception as e:
        logging.error(f"Error in warehouse file upload: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'文件上傳失敗: {str(e)}'}), 500

@app.route('/api/warehouses', methods=['GET'])
def get_warehouses():
    """
    獲取所有倉庫數據
    """
    db: Session = next(get_db())
    try:
        logging.info("Fetching warehouse data from database...")
        warehouses = db.query(Warehouse).all()
        logging.info(f"Found {len(warehouses)} warehouse records")
        
        result = [{
            'warehouseName': w.warehouse_name,
            'productName': w.product_name,
            'quantity': w.quantity,
            'updatedAt': w.updated_at.isoformat() if w.updated_at else None
        } for w in warehouses]
        
        logging.info(f"Returning {len(result)} warehouse records")
        return jsonify(result)
    except Exception as e:
        logging.error(f"Error getting warehouses: {e}", exc_info=True)
        return jsonify({"error": "Could not retrieve warehouse data"}), 500
    finally:
        db.close()

@app.route('/api/warehouse-replenishment-suggestion/<string:store_key>', methods=['POST'])
def get_warehouse_replenishment_suggestion(store_key):
    """
    新的補貨建議 API，專門用於補貨分頁，考慮倉庫庫存
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "No data provided"}), 400

    selected_warehouses = data.get("warehouses", [])
    if not selected_warehouses:
        return jsonify({
            "success": False,
            "message": "請選擇至少一個倉庫"
        }), 400

    strategy = data.get("strategy", "stable")
    machine_capacity = int(data.get("max_total_qty", 50))
    if machine_capacity < 1 or machine_capacity > 50:
        machine_capacity = 50

    db: Session = next(get_db())
    try:
        store_name, machine_id = store_key.split('-', 1)
        
        # 1. 獲取倉庫庫存數據
        warehouse_inventory = {}
        for warehouse_name in selected_warehouses:
            warehouse_items = db.query(Warehouse).filter_by(warehouse_name=warehouse_name).all()
            for item in warehouse_items:
                if item.product_name in warehouse_inventory:
                    warehouse_inventory[item.product_name] += item.quantity
                else:
                    warehouse_inventory[item.product_name] = item.quantity

        if not warehouse_inventory:
            return jsonify({
                "success": False,
                "message": "選擇的倉庫中沒有可用庫存"
            }), 400

        # 2. 獲取機台當前庫存
        inventory_items = db.query(Inventory).filter_by(store=store_name, machine_id=machine_id).all()
        current_inventory = {item.product_name: item.quantity for item in inventory_items}

        # 3. 獲取銷售數據
        thirty_days_ago = datetime.now() - timedelta(days=30)
        transactions = db.query(Transaction).filter(
            Transaction.store_key.startswith(store_name),
            Transaction.transaction_time >= thirty_days_ago
        ).all()

        sales_counts = {}
        for t in transactions:
            if t.product_name:
                sales_counts[t.product_name] = sales_counts.get(t.product_name, 0) + 1

        # 4. 根據策略生成建議
        suggestion_list = []
        warning = None

        # 根據不同策略生成建議
        if strategy == 'stable':
            # 穩健策略：根據銷量比例分配剩餘空間
            sales_products = []
            current_total = sum(current_inventory.values())  # 當前總數
            remaining_space = machine_capacity - current_total  # 剩餘可用空間
            
            for product, warehouse_qty in warehouse_inventory.items():
                current_qty = current_inventory.get(product, 0)
                sales = sales_counts.get(product, 0)
                
                if warehouse_qty > 0:  # 只考慮倉庫有庫存的產品
                    sales_products.append({
                        'productName': product,
                        'currentQty': current_qty,
                        'warehouseQty': warehouse_qty,
                        'salesCount30d': sales,
                        'suggestedQty': current_qty  # 初始設為當前數量
                    })
            
            if sales_products and remaining_space > 0:
                # 根據銷量計算額外分配
                total_sales = sum(p['salesCount30d'] for p in sales_products)
                products_to_add = [p for p in sales_products if p['warehouseQty'] > 0]
                
                if total_sales > 0 and products_to_add:
                    # 有銷量的產品按比例分配剩餘空間
                    base_additions = []
                    for product in products_to_add:
                        ratio = product['salesCount30d'] / total_sales
                        addition = round(ratio * remaining_space)
                        base_additions.append((product, addition))
                        
                    # 調整以確保總數正確
                    total_addition = sum(addition for _, addition in base_additions)
                    if total_addition != remaining_space:
                        diff = remaining_space - total_addition
                        # 按比例調整差異
                        for i, (product, _) in enumerate(base_additions):
                            if i < abs(diff):
                                base_additions[i] = (product, base_additions[i][1] + (1 if diff > 0 else -1))
                    
                    # 應用調整後的數量
                    for product, addition in base_additions:
                        product['suggestedQty'] = product['currentQty'] + addition
                else:
                    # 無銷量時平均分配剩餘空間
                    base_qty = remaining_space // len(products_to_add)
                    remainder = remaining_space % len(products_to_add)
                    for i, product in enumerate(products_to_add):
                        addition = base_qty + (1 if i < remainder else 0)
                        product['suggestedQty'] = product['currentQty'] + addition
                
                suggestion_list.extend(sales_products)

        elif strategy == 'aggressive':
            # 積極策略：優先分配給熱銷品，確保總量為50
            current_total = sum(current_inventory.values())
            remaining_space = machine_capacity - current_total
            
            # 先加入所有現有產品
            suggestion_list = [{
                'productName': product,
                'currentQty': qty,
                'suggestedQty': qty,  # 初始設為當前數量
                'warehouseQty': warehouse_inventory.get(product, 0),
                'salesCount30d': sales_counts.get(product, 0)
            } for product, qty in current_inventory.items()]
            
            # 根據銷量排序所有可能的新增產品
            available_products = [(p, warehouse_inventory.get(p, 0), sales_counts.get(p, 0))
                                for p in warehouse_inventory.keys()
                                if p not in current_inventory and warehouse_inventory.get(p, 0) > 0]
            
            sorted_products = sorted(available_products,
                                  key=lambda x: x[2],  # 按銷量排序
                                  reverse=True)
            
            if remaining_space > 0 and sorted_products:
                # 新產品中的前3名分配80%的剩餘空間
                top_3 = sorted_products[:3]
                top_3_space = round(remaining_space * 0.8)
                base_qty = top_3_space // len(top_3)
                remainder = top_3_space % len(top_3)
                
                for i, (product, warehouse_qty, sales) in enumerate(top_3):
                    suggested_qty = base_qty + (1 if i < remainder else 0)
                    suggestion_list.append({
                        'productName': product,
                        'currentQty': 0,
                        'suggestedQty': suggested_qty,
                        'warehouseQty': warehouse_qty,
                        'salesCount30d': sales
                    })
            
            # 剩餘產品分配剩餘空間
            other_products = sorted_products[3:]
            remaining_space_for_others = remaining_space - top_3_space
            
            if other_products and remaining_space_for_others > 0:
                base_qty = remaining_space_for_others // len(other_products)
                remainder = remaining_space_for_others % len(other_products)
                
                for i, (product, warehouse_qty, sales) in enumerate(other_products):
                    suggested_qty = base_qty + (1 if i < remainder else 0)
                    suggestion_list.append({
                        'productName': product,
                        'currentQty': 0,
                        'suggestedQty': suggested_qty,
                        'warehouseQty': warehouse_qty,
                        'salesCount30d': sales
                    })

        else:  # exploratory
            # 探索策略：80%空間給現有產品，20%給新產品
            suggestion_list = []
            current_total = sum(current_inventory.values())
            remaining_space = machine_capacity - current_total
            
            # 先加入所有現有產品
            for product, qty in current_inventory.items():
                suggestion_list.append({
                    'productName': product,
                    'currentQty': qty,
                    'suggestedQty': qty,  # 保持當前數量
                    'warehouseQty': warehouse_inventory.get(product, 0),
                    'salesCount30d': sales_counts.get(product, 0)
                })
            
            if remaining_space > 0:
                # 處理有銷量但不在當前庫存的產品（佔剩餘空間的80%）
                existing_space = round(remaining_space * 0.8)
                sales_products = [(p, warehouse_inventory[p], sales_counts.get(p, 0))
                                for p in warehouse_inventory.keys()
                                if p not in current_inventory and sales_counts.get(p, 0) > 0 and warehouse_inventory[p] > 0]
                
                if sales_products:
                    base_qty = existing_space // len(sales_products)
                    remainder = existing_space % len(sales_products)
                    
                    for i, (product, warehouse_qty, sales) in enumerate(sales_products):
                        suggested_qty = base_qty + (1 if i < remainder else 0)
                        suggestion_list.append({
                            'productName': product,
                            'currentQty': 0,
                            'suggestedQty': suggested_qty,
                            'warehouseQty': warehouse_qty,
                            'salesCount30d': sales
                        })
            
            # 處理新產品（沒有銷量的產品）
            new_space = remaining_space - existing_space  # 剩餘20%空間給新產品
            new_products = [(p, warehouse_inventory[p])
                          for p in warehouse_inventory.keys()
                          if p not in current_inventory and p not in [x['productName'] for x in suggestion_list] 
                          and warehouse_inventory[p] > 0]
            
            if new_products and new_space > 0:
                base_qty = new_space // len(new_products)
                remainder = new_space % len(new_products)
                
                for i, (product, warehouse_qty) in enumerate(new_products):
                    suggested_qty = base_qty + (1 if i < remainder else 0)
                    suggestion_list.append({
                        'productName': product,
                        'currentQty': 0,
                        'suggestedQty': suggested_qty,
                        'warehouseQty': warehouse_qty,
                        'salesCount30d': 0
                    })        # 5. 排序並返回結果
        suggestion_list.sort(key=lambda x: (x['suggestedQty'] - x['currentQty']), reverse=True)
        
        return jsonify({
            "success": True,
            "store_key": store_key,
            "strategy_used": strategy,
            "suggestion": suggestion_list,
            "warning": warning,
            "warehouse_info": [{
                "name": w,
                "products": len([p for p in warehouse_inventory if w in selected_warehouses])
            } for w in selected_warehouses]
        })

    except Exception as e:
        logging.error(f"Error generating warehouse replenishment suggestion: {e}", exc_info=True)
        return jsonify({
            "success": False,
            "message": f"生成補貨建議時發生錯誤: {str(e)}"
        }), 500
    finally:
        db.close()

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    """
    Returns all transactions in the format expected by presentation.html.
    """
    db: Session = next(get_db())
    try:
        # If a user is logged in, restrict transactions to user's assigned stores
        user_id = session.get('user_id')
        transactions_query = db.query(Transaction)
        if user_id:
            user = db.query(User).filter(User.id == int(user_id)).first()
            if user:
                user_store_keys = [s.store_key for s in user.stores]
                # build filters similar to inventory matching
                trans_filters = []
                for sk in user_store_keys:
                    if '-' in sk:
                        left = sk.rsplit('-', 1)[0]
                        trans_filters.append(Transaction.store_key == sk)
                        trans_filters.append(Transaction.store_key.ilike(f"{left}%"))
                        trans_filters.append(Transaction.store_key.ilike(f"%{left}%"))
                    else:
                        trans_filters.append(Transaction.store_key.ilike(sk))
                        trans_filters.append(Transaction.store_key.ilike(f"%{sk}%"))
                if trans_filters:
                    transactions_query = transactions_query.filter(or_(*trans_filters))

        transactions = transactions_query.all()

        # Create a map of store_key to store_name for quick lookup
        stores = {s.store_key: s.store_key.rsplit('-', 1)[0] for s in db.query(Store).all()}

        result = [
            {
                "shopName": stores.get(t.store_key, t.store_key), # Fallback to store_key if not in map
                "date": t.transaction_time.isoformat(),
                "amount": t.amount,
                "product": t.product_name,
                "payType": t.payment_type,
            }
            for t in transactions
        ]
        return jsonify(result)
    except Exception as e:
        logging.error(f"Error getting transactions: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        db.close()


@app.route('/api/inventory/<store_key>', methods=['DELETE'])
def delete_inventory_data(store_key):
    """
    Deletes all inventory and custom data associated with a specific store_key.
    """
    if not store_key:
        return jsonify({"success": False, "message": "store_key is required"}), 400

    db: Session = next(get_db())
    try:
        store_name, machine_id = store_key.split('-', 1)

        db.query(Inventory).filter_by(store=store_name, machine_id=machine_id).delete(synchronize_session=False)
        db.query(Store).filter_by(store_key=store_key).delete(synchronize_session=False)
        db.query(Transaction).filter_by(store_key=store_key).delete(synchronize_session=False)

        db.commit()

        return jsonify({"success": True, "message": "Data deleted successfully."})

    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        db.close()


def distribute_remainder(items, total_slots):
    """
    一個輔助函數，用於處理補貨建議數量計算中的小數問題。
    它會確保所有產品的建議數量加總後剛好等於機台的目標總容量。
    """
    # 根據小數部分由大到小排序，小數越大的越優先獲得 +1
    items.sort(key=lambda x: x['suggestedQty_float'] - int(x['suggestedQty_float']), reverse=True)
    
    # 計算所有品項無條件捨去後的總和
    current_total = sum(int(item['suggestedQty_float']) for item in items)
    remainder = total_slots - current_total
    
    result = []
    for i, item in enumerate(items):
        qty = int(item['suggestedQty_float'])
        # 將餘下的數量逐一分配給排序最前面的品項
        if i < remainder:
            qty += 1
        result.append({'productName': item['productName'], 'suggestedQty': qty, 'sales_count': item.get('sales_count', 0)})
    
    return result

@app.route('/api/replenishment-suggestion/<string:store_key>', methods=['POST'])
def get_replenishment_suggestion(store_key):
    """
    生成補貨建議的核心API。
    接收策略和預留空位，回傳一份詳細的補貨清單。
    """
    data = request.get_json()
    if not data:
        return jsonify({"success": False, "message": "No data provided"}), 400

    strategy = data.get("strategy", "stable")
    reserve_slots = int(data.get("reserve_slots", 0))
    only_add = bool(data.get("only_add", False))
    machine_capacity = int(data.get("max_total_qty", 50))
    selected_warehouses = data.get("warehouses", [])
    
    if not selected_warehouses:
        return jsonify({
            "success": False,
            "message": "請選擇至少一個倉庫"
        }), 400

    if machine_capacity < 1 or machine_capacity > 50:
        machine_capacity = 50
    
    available_slots = machine_capacity - reserve_slots

    db: Session = next(get_db())
    try:
        store_name, machine_id = store_key.split('-', 1)
        
        # 1. 獲取目前庫存
        inventory_items = db.query(Inventory).filter_by(store=store_name, machine_id=machine_id).all()
        current_inventory = {item.product_name: item.quantity for item in inventory_items}

        # 2. 獲取過去30天的銷售數據 (此處假設同店鋪名的所有機台共享銷售數據)
        thirty_days_ago = datetime.now() - timedelta(days=30)
        transactions = db.query(Transaction).filter(
            Transaction.store_key.startswith(store_name),
            Transaction.transaction_time >= thirty_days_ago
        ).all()

        sales_counts = {}
        for t in transactions:
            if t.product_name:
                sales_counts[t.product_name] = sales_counts.get(t.product_name, 0) + 1
        
        total_sales_volume = sum(sales_counts.values())

        # 2.5 獲取所選倉庫的庫存數據
        warehouse_inventory = {}
        for warehouse_name in selected_warehouses:
            warehouse_items = db.query(Warehouse).filter_by(warehouse_name=warehouse_name).all()
            for item in warehouse_items:
                if item.product_name in warehouse_inventory:
                    warehouse_inventory[item.product_name] += item.quantity
                else:
                    warehouse_inventory[item.product_name] = item.quantity

        if not warehouse_inventory:
            return jsonify({
                "success": False,
                "message": "選擇的倉庫中沒有可用庫存"
            }), 400

        if total_sales_volume == 0:
            # 如果沒有銷售數據，則根據倉庫庫存情況提供建議
            available_products = list(warehouse_inventory.items())
            if strategy == 'stable':
                # 平均分配倉庫中有的商品
                slots_per_product = available_slots // len(available_products)
                suggestion_list = [
                    {'productName': p, 'suggestedQty': min(slots_per_product, q)} 
                    for p, q in available_products
                ]
            elif strategy == 'aggressive':
                # 按倉庫庫存量排序，優先分配庫存量大的商品
                sorted_products = sorted(available_products, key=lambda x: x[1], reverse=True)
                top_products = sorted_products[:3]
                suggestion_list = [
                    {'productName': p, 'suggestedQty': min(round(available_slots * 0.3), q)} 
                    for p, q in top_products
                ]
            else:  # exploratory
                # 少量嘗試倉庫中的所有商品
                suggestion_list = [
                    {'productName': p, 'suggestedQty': min(2, q)} 
                    for p, q in available_products
                ]
            
            return jsonify({
                "success": True,
                "strategy_used": f"{strategy}_no_sales",
                "suggestion": suggestion_list,
                "message": "根據倉庫庫存生成建議"
            })

        # 3. 應用不同策略（考慮銷售數據和倉庫庫存）
        suggestion_list = []
        # 只考慮倉庫中有庫存的產品的銷售數據
        filtered_sales = {
            product: count for product, count in sales_counts.items() 
            if product in warehouse_inventory
        }
        if not filtered_sales:
            return jsonify({
                "success": False,
                "message": "倉庫中沒有任何有銷售記錄的產品"
            }), 400
            
        total_filtered_sales = sum(filtered_sales.values())
        sorted_sales = sorted(filtered_sales.items(), key=lambda item: item[1], reverse=True)

        if strategy == 'stable':
            # 穩健策略：根據銷售比例分配，但受倉庫庫存限制
            temp_suggestions = []
            for p, c in filtered_sales.items():
                suggested_qty = (c / total_filtered_sales) * available_slots
                # 確保不超過倉庫庫存
                warehouse_qty = warehouse_inventory.get(p, 0)
                suggested_qty = min(suggested_qty, warehouse_qty)
                temp_suggestions.append({
                    'productName': p,
                    'suggestedQty_float': suggested_qty,
                    'sales_count': c,
                    'warehouse_qty': warehouse_qty
                })
            suggestion_list = distribute_remainder(temp_suggestions, available_slots)
        
        elif strategy == 'aggressive':
            # 積極策略：優先分配銷量前三的產品，但受倉庫庫存限制
            top_3_products = sorted_sales[:3]
            other_products = sorted_sales[3:]
            
            slots_for_top_3 = round(available_slots * 0.8)
            slots_for_others = available_slots - slots_for_top_3
            
            temp_suggestions = []
            # 處理前三名產品
            for p, c in top_3_products:
                warehouse_qty = warehouse_inventory.get(p, 0)
                suggested_qty = min(slots_for_top_3 / len(top_3_products), warehouse_qty)
                temp_suggestions.append({
                    'productName': p,
                    'suggestedQty_float': suggested_qty,
                    'sales_count': c,
                    'warehouse_qty': warehouse_qty
                })
            
            # 處理其他產品
            if other_products:
                qty_per_other = slots_for_others / len(other_products)
                for p, c in other_products:
                    warehouse_qty = warehouse_inventory.get(p, 0)
                    suggested_qty = min(qty_per_other, warehouse_qty)
                    temp_suggestions.append({
                        'productName': p,
                        'suggestedQty_float': suggested_qty,
                        'sales_count': c,
                        'warehouse_qty': warehouse_qty
                    })
            
            suggestion_list = distribute_remainder(temp_suggestions, available_slots)

        elif strategy == 'exploratory':
            # 探索策略：保留20%空間給新產品，其餘根據銷量分配
            slots_for_existing = round(available_slots * 0.8)
            slots_for_new = available_slots - slots_for_existing

            # 處理現有產品
            temp_suggestions = []
            for p, c in filtered_sales.items():
                warehouse_qty = warehouse_inventory.get(p, 0)
                suggested_qty = min(
                    (c / total_filtered_sales) * slots_for_existing,
                    warehouse_qty
                )
                temp_suggestions.append({
                    'productName': p,
                    'suggestedQty_float': suggested_qty,
                    'sales_count': c,
                    'warehouse_qty': warehouse_qty
                })

            # 尋找倉庫中有庫存但尚未銷售的新產品
            new_products = [
                p for p in warehouse_inventory.keys()
                if p not in filtered_sales and warehouse_inventory[p] > 0
            ]
            
            # 為新產品分配空間
            if new_products:
                slots_per_new = slots_for_new / len(new_products)
                for p in new_products:
                    warehouse_qty = warehouse_inventory[p]
                    suggested_qty = min(slots_per_new, warehouse_qty)
                    temp_suggestions.append({
                        'productName': p,
                        'suggestedQty_float': suggested_qty,
                        'sales_count': 0,
                        'warehouse_qty': warehouse_qty
                    })
            
            suggestion_list = distribute_remainder(temp_suggestions, available_slots)
            temp_suggestions = [{'productName': p, 'suggestedQty_float': (c / total_sales_volume) * slots_for_existing, 'sales_count': c} for p, c in sales_counts.items()]
            suggestion_list = distribute_remainder(temp_suggestions, slots_for_existing)

        # 4. 組合最終結果
        final_suggestion = []
        suggestion_map = {item['productName']: item['suggestedQty'] for item in suggestion_list}
        all_product_names = set(current_inventory.keys()) | set(suggestion_map.keys())

        # 正確邏輯：最大補貨總數量是補貨後的目標庫存量
        warning = None
        total_current = sum(current_inventory.get(name, 0) for name in all_product_names)
        
        # 先將所有產品的建議數量設為不低於當前庫存
        for name in all_product_names:
            current_qty = current_inventory.get(name, 0)
            suggested_qty = max(current_qty, suggestion_map.get(name, current_qty))
            final_suggestion.append({
                'productName': name,
                'currentQty': current_qty,
                'suggestedQty': suggested_qty,
                'salesCount30d': sales_counts.get(name, 0)
            })

        # 按照調整量（建議數量-當前數量）排序，只保留前7個需要調整的項目
        final_suggestion.sort(key=lambda x: (
            x['suggestedQty'] - x['currentQty'],  # 首要條件：調整量
            x['salesCount30d']  # 次要條件：銷量
        ), reverse=True)

        # 只保留需要調整的前7個項目，其他項目的建議數量設為當前庫存
        needs_adjustment = [x for x in final_suggestion if x['suggestedQty'] - x['currentQty'] > 0]
        no_adjustment = [x for x in final_suggestion if x['suggestedQty'] - x['currentQty'] <= 0]

        if len(needs_adjustment) > 7:
            for item in needs_adjustment[7:]:
                item['suggestedQty'] = item['currentQty']

        final_suggestion = needs_adjustment[:7] + no_adjustment
        final_suggestion.sort(key=lambda x: x['suggestedQty'], reverse=True)

        if total_current >= machine_capacity:
            warning = f"現有庫存總和({total_current})已達最大補貨總數量({machine_capacity})，無需補貨。"
            for name in all_product_names:
                final_suggestion.append({
                    'productName': name,
                    'currentQty': current_inventory.get(name, 0),
                    'suggestedQty': current_inventory.get(name, 0),
                    'salesCount30d': sales_counts.get(name, 0)
                })
            final_suggestion.sort(key=lambda x: x['suggestedQty'], reverse=True)
            return jsonify({
                "success": True,
                "store_key": store_key,
                "strategy_used": strategy,
                "suggestion": final_suggestion,
                "warning": warning
            })

        # 剩餘可補貨空間
        available_replenish = machine_capacity - total_current
        # 依策略分配這些空間
        # 重新計算分配：每個產品的補貨量 = 分配量，建議數量 = 現有庫存 + 分配量
        # 先依照策略分配比例
        # 取得有銷量的產品分配比例
        if strategy == 'stable':
            total_sales_volume = sum(sales_counts.values())
            temp_suggestions = []
            for p, c in sales_counts.items():
                temp_suggestions.append({'productName': p, 'suggestedQty_float': (c / total_sales_volume) * available_replenish, 'sales_count': c})
            distributed = distribute_remainder(temp_suggestions, available_replenish)
        elif strategy == 'aggressive':
            sorted_sales = sorted(sales_counts.items(), key=lambda item: item[1], reverse=True)
            top_3_products = sorted_sales[:3]
            other_products = sorted_sales[3:]
            top_3_sales_volume = sum(c for _, c in top_3_products)
            other_sales_volume = sum(c for _, c in other_products)
            slots_for_top_3 = round(available_replenish * 0.8)
            slots_for_others = available_replenish - slots_for_top_3
            temp_suggestions = []
            if top_3_sales_volume > 0:
                temp_suggestions.extend([{'productName': p, 'suggestedQty_float': (c / top_3_sales_volume) * slots_for_top_3, 'sales_count': c} for p, c in top_3_products])
            if other_sales_volume > 0 and len(other_products) > 0:
                temp_suggestions.extend([{'productName': p, 'suggestedQty_float': (c / other_sales_volume) * slots_for_others, 'sales_count': c} for p, c in other_products])
            distributed = distribute_remainder(temp_suggestions, available_replenish)
        elif strategy == 'exploratory':
            total_sales_volume = sum(sales_counts.values())
            slots_for_existing = round(available_replenish * 0.8)
            temp_suggestions = [{'productName': p, 'suggestedQty_float': (c / total_sales_volume) * slots_for_existing, 'sales_count': c} for p, c in sales_counts.items()]
            distributed = distribute_remainder(temp_suggestions, slots_for_existing)
        else:
            distributed = []

        # 將分配結果轉為 dict
        distributed_map = {item['productName']: item['suggestedQty'] for item in distributed}

        # 組合最終建議：建議數量 = 現有庫存 + 分配到的補貨量
        for name in all_product_names:
            current_qty = current_inventory.get(name, 0)
            add_qty = distributed_map.get(name, 0)
            suggested_qty = current_qty + add_qty
            # 只補貨模式下，不建議減少現有庫存
            if only_add and suggested_qty < current_qty:
                suggested_qty = current_qty
            final_suggestion.append({
                'productName': name,
                'currentQty': current_qty,
                'suggestedQty': suggested_qty,
                'salesCount30d': sales_counts.get(name, 0)
            })
        final_suggestion.sort(key=lambda x: x['suggestedQty'], reverse=True)

        # 最終檢查總和，理論上不會超過 machine_capacity
        total_final = sum(item['suggestedQty'] for item in final_suggestion)
        if total_final > machine_capacity:
            warning = f"分配後總庫存({total_final})超過最大補貨總數量({machine_capacity})，已自動調整至上限。"
            # 依現有庫存排序，依序減少至符合上限
            over = total_final - machine_capacity
            for item in sorted(final_suggestion, key=lambda x: x['suggestedQty'], reverse=True):
                if over <= 0:
                    break
                reducible = item['suggestedQty'] - item['currentQty']
                if reducible > 0:
                    reduce_by = min(reducible, over)
                    item['suggestedQty'] -= reduce_by
                    over -= reduce_by
            # 再次排序
            final_suggestion.sort(key=lambda x: x['suggestedQty'], reverse=True)

        return jsonify({
            "success": True,
            "store_key": store_key,
            "strategy_used": strategy,
            "suggestion": final_suggestion,
            "warning": warning
        })

    except Exception as e:
        db.rollback()
        logging.error(f"Error in replenishment suggestion for {store_key}: {e}", exc_info=True)
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        db.close()

# --- Password Protected Routes ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handles the login process."""
    if request.method == 'POST':
        password = request.form.get('password')
        # Load the access password from environment variables
        access_password = os.getenv("ACCESS_PASSWORD")
        
        if not access_password:
             flash("錯誤：應用程式未設定訪問密碼。")
             return render_template('login.html'), 500

        if password == access_password:
            session['logged_in'] = True
            return redirect('/')
        else:
            flash('密碼錯誤，請重試。')
            return redirect('/login')
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    """Logs the user out."""
    session.pop('logged_in', None)
    return redirect('/login')


# --- Static File Serving ---
@app.route('/')
def serve_index():
    if not session.get('logged_in'):
        return redirect('/login')
    return send_from_directory(script_dir, 'index.html')

@app.route('/presentation')
def serve_presentation():
    return send_from_directory(script_dir, 'presentation.html')


@app.route('/presentation_sample')
def serve_presentation_sample():
    # Legacy route: redirect to the canonical sample path
    return redirect('/sample')


@app.route('/sample')
def serve_sample():
    # Canonical sample URL (no .html)
    return send_from_directory(script_dir, 'presentation_sample.html')

@app.route('/<path:path>')
def serve_static_files(path):
    # This will serve other files like script.js
    # Allow presentation_sample.html to be requested directly but prefer clean URL
    if path == 'presentation.html':
        return redirect('/presentation')
    if path == 'presentation_sample.html':
        return redirect('/sample')
    return send_from_directory(script_dir, path)


# --- Scheduler Setup ---
def run_scheduler():
    """
    Sets up and runs the scheduler in a loop.
    """
    # Schedule the inventory scraper to run at 1 minute past the hour.
    schedule.every().hour.at(":01").do(run_inventory_scraper_background)
    logging.info("Scheduler started for inventory: will run every hour at 1 minute past.")

    # Schedule the sales scraper to run daily at 23:55 Taiwan Time (UTC+8), which is 15:55 UTC.
    schedule.every().day.at("15:55").do(run_sales_scraper_background)
    logging.info("Scheduler started for sales: will run daily at 15:55 UTC (23:55 Taiwan Time).")
    
    # Schedule the warehouse scraper to run daily at 23:50 Taiwan Time (UTC+8), which is 15:50 UTC.
    schedule.every().day.at("15:50").do(run_warehouse_scraper_background)
    logging.info("Scheduler started for warehouse: will run daily at 15:50 UTC (23:50 Taiwan Time).")
    
    # Run the scheduler loop
    while True:
        schedule.run_pending()
        time.sleep(1)


@app.route('/debug/db-stats')
def debug_db_stats():
    """Temporary diagnostic endpoint: returns counts of key tables and DB config."""
    # WARNING: This endpoint may expose counts of rows; keep it temporary and remove after debugging.
    db: Session = next(get_db())
    try:
        inv_count = db.query(Inventory).count()
        wh_count = db.query(Warehouse).count()
        store_count = db.query(Store).count()
        return jsonify({
            "success": True,
            "inventory_count": inv_count,
            "warehouse_count": wh_count,
            "store_count": store_count,
            "uses_external_database": bool(os.getenv('DATABASE_URL'))
        })
    except Exception as e:
        logging.exception('Error while gathering DB stats')
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        db.close()

# --- Endpoints for Manual Testing ---
@app.route('/test-run-warehouse-scraper', methods=['GET'])
def test_run_warehouse_scraper():
    """用於手動觸發倉庫爬蟲的測試端點。"""
    logging.info("收到手動觸發倉庫爬蟲的請求。")
    thread = threading.Thread(target=run_warehouse_scraper_background)
    thread.start()
    return "倉庫爬蟲工作已手動觸發進行測試。"

@app.route('/test-run-inventory-scraper', methods=['GET'])
def test_run_inventory_scraper():
    """A simple endpoint to manually trigger the inventory scraper for testing."""
    logging.info("Manual trigger for inventory scraper received.")
    thread = threading.Thread(target=run_inventory_scraper_background)
    thread.start()
    return "Inventory scraper job manually triggered for testing."

@app.route('/test-run-sales-scraper', methods=['GET'])
def test_run_sales_scraper():
    """A simple endpoint to manually trigger the sales scraper for testing."""
    logging.info("Manual trigger for sales scraper received.")
    thread = threading.Thread(target=run_sales_scraper_background)
    thread.start()
    return "Sales scraper job manually triggered for testing."


# --- 補貨單生成相關功能 ---
@app.route('/api/generate-replenishment-form', methods=['POST'])
def generate_replenishment_form():
    """根據補貨建議生成補貨單Excel"""
    try:
        # 獲取請求數據
        data = request.get_json()
        if not data or 'suggestions' not in data:
            return jsonify({'success': False, 'message': '無效的請求數據'}), 400
        
        suggestions = data['suggestions']
        if not suggestions:
            return jsonify({'success': False, 'message': '沒有補貨建議數據'}), 400

        # 加載模板
        template_path = os.path.join(script_dir, '補貨單.xlsx')
        current_date = datetime.now().strftime('%Y%m%d')
        output_path = os.path.join(script_dir, f'{current_date}-在那裡-每日商品補貨明細表.xlsx')
        
        if not os.path.exists(template_path):
            return jsonify({'success': False, 'message': '找不到補貨單模板'}), 404
            
        # 複製模板
        shutil.copy2(template_path, output_path)
        
        # 打開工作簿
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 填寫基本信息
        ws['F1'] = datetime.now().strftime('%Y/%m/%d')  # 補貨日期
        ws['H2'] = len(suggestions)  # 增加點位數
        
        # 機台位置映射
        machine_positions = [
            {'name': 'A4', 'items': ('B4', 'D4', 'B12', 'D12'), 'item_range': (4, 11)},  # 第一台
            {'name': 'E4', 'items': ('F4', 'H4', 'F12', 'H12'), 'item_range': (4, 11)},  # 第二台
            {'name': 'A17', 'items': ('B17', 'D17', 'B25', 'D25'), 'item_range': (17, 24)},  # 第三台
            {'name': 'E17', 'items': ('F17', 'H17', 'F25', 'H25'), 'item_range': (17, 24)},  # 第四台
            {'name': 'A30', 'items': ('B30', 'D30', 'B38', 'D38'), 'item_range': (30, 37)},  # 第五台
            {'name': 'E30', 'items': ('F30', 'H30', 'F38', 'H38'), 'item_range': (30, 37)},  # 第六台
            {'name': 'A43', 'items': ('B43', 'D43', 'B51', 'D51'), 'item_range': (43, 50)},  # 第七台
            {'name': 'E43', 'items': ('F43', 'H43', 'E51', 'H51'), 'item_range': (43, 50)}   # 第八台
        ]

        # 填寫每台機器的數據
        for i, suggestion in enumerate(suggestions[:8]):  # 最多處理8台機器
            pos = machine_positions[i]
            ws[pos['name']] = suggestion['machine']  # 填寫機台名稱
            
            # 獲取需要增加庫存的產品（只取正數調整）
            adjusted_items = [item for item in suggestion['suggestion'] 
                            if item['suggestedQty'] - item['currentQty'] > 0]
            
            # 按調整數量排序並只取前7個
            adjusted_items.sort(key=lambda x: x['suggestedQty'] - x['currentQty'], reverse=True)
            adjusted_items = adjusted_items[:7]  # 限制每台最多7個調整項目
            
            # 填寫產品信息
            for j, item in enumerate(adjusted_items):
                row = pos['item_range'][0] + j
                item_col = pos['items'][0][0]  # B或F
                qty_col = pos['items'][1][0]   # D或H
                
                # 填寫產品名稱和調整數量
                ws[f'{item_col}{row}'] = item['productName']
                ws[f'{qty_col}{row}'] = item['suggestedQty'] - item['currentQty']
            
            # 填寫庫存和總數
            current_stock = sum(item['currentQty'] for item in suggestion['suggestion'])
            total_adjustment = sum(item['suggestedQty'] - item['currentQty'] 
                                 for item in suggestion['suggestion'])
            
            # 使用machine_positions中定義的單元格位置
            ws[pos['items'][2]] = current_stock  # 機內庫存
            ws[pos['items'][3]] = total_adjustment  # 補貨總數

        # 保存工作簿
        wb.save(output_path)
        
        # 返回文件名供下載
        filename = os.path.basename(output_path)
        return jsonify({
            'success': True, 
            'message': '補貨單生成成功',
            'filename': filename
        })
        
    except Exception as e:
        logging.error(f"生成補貨單時發生錯誤: {str(e)}")
        return jsonify({'success': False, 'message': f'生成補貨單時發生錯誤: {str(e)}'}), 500

@app.route('/download-replenishment-form/<filename>')
def download_replenishment_form(filename):
    """下載生成的補貨單"""
    try:
        return send_from_directory(script_dir, filename, as_attachment=True)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 404

# --- 銷售明細表生成相關功能 ---
@app.route('/api/generate-sales-detail', methods=['POST'])
def generate_sales_detail():
    """根據日期範圍生成銷售明細表Excel"""
    try:
        data = request.get_json()
        if not data or 'startDate' not in data or 'endDate' not in data:
            return jsonify({'success': False, 'message': '無效的請求數據'}), 400

        start_date = datetime.fromisoformat(data['startDate'])
        end_date = datetime.fromisoformat(data['endDate'])

        # 可選的分店/產品過濾條件（來自前端多選）
        selected_stores = data.get('stores') or []
        selected_products = data.get('products') or []

        logging.info(f"Received generate-sales-detail request with stores={selected_stores} products={selected_products}")

        db: Session = next(get_db())
        try:
            # 設定日期範圍的結束時間為當天的最後一刻
            end_date = end_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            
            # 獲取指定日期範圍內的所有交易
            transactions_query = db.query(Transaction).filter(
                Transaction.transaction_time >= start_date,
                Transaction.transaction_time <= end_date
            )

            # 如果前端有提供分店過濾，將 store_key 的店名部分比對
            if selected_stores:
                # store_key 格式為 'StoreName-machineId'，因此比對以 store_key.startswith(storeName)
                store_filters = [Transaction.store_key.startswith(s) for s in selected_stores]
                # SQLAlchemy 不能直接使用 python list of expressions like this; 使用 OR 組合
                from sqlalchemy import or_
                or_conditions = [Transaction.store_key.startswith(s) for s in selected_stores]
                transactions_query = transactions_query.filter(or_(*or_conditions))

            # 如果前端有提供產品過濾
            if selected_products:
                from sqlalchemy import or_
                prod_conditions = [Transaction.product_name == p for p in selected_products]
                transactions_query = transactions_query.filter(or_(*prod_conditions))

            transactions = transactions_query.order_by(Transaction.store_key, Transaction.product_name).all()

            if not transactions:
                return jsonify({'success': False, 'message': '指定日期範圍內沒有交易記錄'}), 404
                
            # 記錄查詢到的交易記錄範圍
            logging.info(f"Query date range: from {start_date} to {end_date}")
            logging.info(f"Found {len(transactions)} transactions after applying filters (stores/products)")

            # 建立 Excel 工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "銷售明細"

            # 設置標題行
            ws['A1'] = '商品'
            ws['B1'] = '單價'
            ws['C1'] = '份數'
            ws['D1'] = '小計'

            # 統計數據
            sales_summary = {}
            total_amount = 0  # 用於驗證總金額
            transaction_count = 0  # 用於驗證交易筆數
            
            # 第一次遍歷：找出每個產品的最新單價（用於單價回復），
            # 但我們允許同一產品存在多個不同的已還原單價，
            # 因此此階段只負責計算單筆交易的還原單價。
            def normalized_unit_price(amount):
                try:
                    price = int(round(amount))
                except Exception:
                    price = int(amount)
                last_digit = price % 10
                if last_digit in (1, 2):
                    return price - 2
                return price
            
            # 第二次遍歷：計算每個店家的每個產品在不同（還原後）單價下的數量與營收
            # 結構：sales_summary[store][product][price] -> {'count': n, 'total': x}
            sales_summary = {}
            for t in transactions:
                transaction_count += 1
                store_name = t.store_key.split('-')[0]
                product_name = t.product_name or 'UNKNOWN'

                # 每筆交易金額 = 實際交易金額
                transaction_amount = t.amount or 0
                total_amount += transaction_amount

                # 以還原後的單價作為 group key
                unit_price = normalized_unit_price(t.amount)

                store_map = sales_summary.setdefault(store_name, {})
                product_map = store_map.setdefault(product_name, {})
                price_entry = product_map.setdefault(unit_price, {'count': 0, 'total': 0})
                price_entry['count'] += 1
                price_entry['total'] += transaction_amount
            
            # 詳細的日誌記錄
            logging.info(f"Total transactions: {transaction_count}")
            logging.info(f"Total amount: {total_amount}")

            # 計算唯一的 product-price 組合數與驗證總金額
            unique_price_combinations = 0
            summary_total = 0
            for store_name, products in sales_summary.items():
                for product_name, price_map in products.items():
                    for price, rec in price_map.items():
                        unique_price_combinations += 1
                        summary_total += rec.get('total', 0)

            logging.info(f"Unique product-price combinations: {unique_price_combinations}")
            if summary_total != total_amount:
                logging.error(f"Amount mismatch! Summary: {summary_total}, Transactions: {total_amount}")

            # 按店鋪匯總統計以供日誌顯示
            store_totals = {}
            for store_name, products in sales_summary.items():
                store_totals.setdefault(store_name, {'total': 0, 'transactions': 0})
                for product_name, price_map in products.items():
                    for price, rec in price_map.items():
                        store_totals[store_name]['total'] += rec.get('total', 0)
                        store_totals[store_name]['transactions'] += rec.get('count', 0)
            
            for store, stats in store_totals.items():
                logging.info(f"Store: {store}")
                logging.info(f"  - Total amount: {stats['total']}")
                logging.info(f"  - Transaction count: {stats['transactions']}")
                logging.info(f"  - Average transaction: {stats['total'] / stats['transactions']:.2f}")

            # 按店家分組並支援每個產品多個價格（以還原後的單價為鍵）
            # sales_summary 現在以 store->product->price 聚合
            store_groups = {}
            for store_name in sorted(sales_summary.keys()):
                store_entry = {'sales': [], 'total_count': 0, 'total_amount': 0}
                for product_name in sorted(sales_summary[store_name].keys()):
                    price_map = sales_summary[store_name][product_name]
                    # 將不同價格按價格高到低排序
                    price_rows = []
                    for price in sorted(price_map.keys(), reverse=True):
                        rec = price_map[price]
                        price_rows.append({'price': price, 'count': rec['count'], 'total': rec['total']})
                        store_entry['total_count'] += rec['count']
                        store_entry['total_amount'] += rec['total']

                    store_entry['sales'].append({'product': product_name, 'prices': price_rows})

                # 根據店內總收入排序產品（主要顯示順序）
                store_entry['sales'].sort(key=lambda p: sum(r['total'] for r in p['prices']), reverse=True)
                store_groups[store_name] = store_entry

            # 設置樣式
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            # 字體樣式
            header_font = Font(bold=True, size=11, name='微軟正黑體')
            normal_font = Font(size=10, name='微軟正黑體')
            store_font = Font(bold=True, size=11, name='微軟正黑體')  # 調整為與標題相同大小
            
            # 填充顏色
            header_fill = PatternFill(start_color='2F75B5', end_color='2F75B5', fill_type='solid')  # 主題藍
            subtotal_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # 淺藍色
            total_fill = PatternFill(start_color='8EA9DB', end_color='8EA9DB', fill_type='solid')    # 中藍色
            alternate_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid') # 淺灰色
            
            # 邊框樣式
            thin_border = Border(
                left=Side(style='thin', color='BFBFBF'),
                right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'),
                bottom=Side(style='thin', color='BFBFBF')
            )
            
            # 特殊邊框樣式（用於標題和小計行）
            bottom_border = Border(
                left=Side(style='thin', color='BFBFBF'),
                right=Side(style='thin', color='BFBFBF'),
                top=Side(style='thin', color='BFBFBF'),
                bottom=Side(style='medium', color='2F75B5')  # 底部使用較粗的藍色邊框
            )
            
            # 對齊方式
            center_alignment = Alignment(horizontal='center', vertical='center')
            right_alignment = Alignment(horizontal='right', vertical='center')
            
            # 設置標題樣式
            headers = ['商品', '單價', '份數', '小計']  # 移除店鋪欄位
            for col, header in zip(['A1', 'B1', 'C1', 'D1'], headers):
                cell = ws[col]
                cell.value = header
                cell.font = Font(bold=True, size=11, name='微軟正黑體', color='FFFFFF')
                cell.fill = header_fill
                cell.border = bottom_border
                cell.alignment = center_alignment

            # 填入數據，包括每個店家的小計
            current_row = 2
            grand_total_count = 0
            grand_total_amount = 0

            for store, data in store_groups.items():
                # 店家名稱行
                ws.merge_cells(f'A{current_row}:D{current_row}')
                ws[f'A{current_row}'] = store
                ws[f'A{current_row}'].font = store_font
                ws[f'A{current_row}'].fill = subtotal_fill
                ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{current_row}'].border = thin_border
                
                last_header_row = current_row  # 記錄店家標題行的位置
                current_row += 1

                # 填入該店家的所有商品數據（支援同商品多價格）
                # 使用 product_block_idx 來對整個商品區塊進行交替底色，
                # 以避免因為同商品多價格造成的列數差異而導致視覺上跨店家錯位。
                product_block_idx = 0
                for sale in data['sales']:
                    product_name = sale['product']
                    price_rows = sale.get('prices', [])
                    if not price_rows:
                        # 若沒有價格資料，跳過（且不計入交替序列）
                        continue

                    # 決定該商品區塊是否需要填充背景（保持整個商品區塊同色）
                    should_fill_block = (product_block_idx % 2 == 1)

                    start_row_for_product = current_row
                    # 為每個價格列寫入一行
                    for pr in price_rows:
                        ws[f'B{current_row}'] = pr['price']
                        ws[f'C{current_row}'] = pr['count']
                        ws[f'D{current_row}'] = pr['total']

                        # 設定樣式
                        for col, value in [(f'B{current_row}', pr['price']), (f'C{current_row}', pr['count']), (f'D{current_row}', pr['total'])]:
                            ws[col].font = normal_font
                            ws[col].border = thin_border
                            ws[col].alignment = Alignment(horizontal='right', vertical='center')
                            if col.startswith('B'):
                                ws[col].number_format = '#,##0'
                            else:
                                ws[col].number_format = '#,##0'

                        # 如果該商品區塊需要填色，對整行所有列填充
                        if should_fill_block:
                            for col in ['A', 'B', 'C', 'D']:
                                ws[f'{col}{current_row}'].fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

                        current_row += 1

                    end_row_for_product = current_row - 1
                    # 合併 A 欄為產品名稱以提高可讀性（整個商品區塊不另加空行）
                    ws.merge_cells(start_row=start_row_for_product, start_column=1, end_row=end_row_for_product, end_column=1)
                    ws[f'A{start_row_for_product}'] = product_name
                    ws[f'A{start_row_for_product}'].font = normal_font
                    ws[f'A{start_row_for_product}'].alignment = Alignment(horizontal='left', vertical='center')

                    # 商品區塊處理完畢，移動到下一個商品塊
                    product_block_idx += 1

                # 添加該店家的小計
                ws[f'A{current_row}'] = f"{store} 小計"
                ws[f'C{current_row}'] = data['total_count']
                ws[f'D{current_row}'] = data['total_amount']
                
                # 設置小計行的樣式
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{current_row}'].fill = subtotal_fill
                    ws[f'{col}{current_row}'].font = header_font
                    ws[f'{col}{current_row}'].border = thin_border
                    ws[f'{col}{current_row}'].alignment = right_alignment
                
                # 為小計的數字設置格式
                ws[f'C{current_row}'].number_format = '#,##0'  # 總份數使用千分位
                ws[f'D{current_row}'].number_format = '#,##0'  # 總金額使用千分位
                ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')

                current_row += 2  # 添加一個空行

                # 累加到總計
                grand_total_count += data['total_count']
                grand_total_amount += data['total_amount']

            current_row -= 1  # 移除最後一個多餘的空行

            # 添加總計行
            ws[f'A{current_row}'] = "總計"
            ws[f'C{current_row}'] = grand_total_count
            ws[f'D{current_row}'] = grand_total_amount

            # 設置總計行的樣式
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{current_row}'].fill = total_fill
                ws[f'{col}{current_row}'].font = header_font
                ws[f'{col}{current_row}'].border = thin_border
                ws[f'{col}{current_row}'].alignment = right_alignment
            
            # 為總計的數字設置格式
            ws[f'C{current_row}'].number_format = '#,##0'  # 總份數使用千分位
            ws[f'D{current_row}'].number_format = '#,##0'  # 總金額使用千分位
            ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')

            # 設置欄寬
            ws.column_dimensions['A'].width = 45  # 商品名稱（增加寬度）
            ws.column_dimensions['B'].width = 15  # 單價
            ws.column_dimensions['C'].width = 12  # 份數
            ws.column_dimensions['D'].width = 18  # 小計

            # 設置列高
            ws.row_dimensions[1].height = 25  # 標題列加高
            
            # 設置工作表其他屬性
            ws.freeze_panes = 'A2'  # 凍結首行
            ws.sheet_view.showGridLines = False  # 隱藏網格線

            # 生成日期範圍字串
            start_str = start_date.strftime('%Y%m%d')
            end_str = end_date.strftime('%Y%m%d')
            filename = f'sales_detail_{start_str}-{end_str}.xlsx'
            output_path = os.path.join(script_dir, filename)
            wb.save(output_path)

            return jsonify({
                'success': True,
                'message': '明細表生成成功',
                'filename': filename
            })

        except Exception as e:
            logging.error(f"生成明細表時發生錯誤: {str(e)}", exc_info=True)
            return jsonify({'success': False, 'message': f'生成明細表時發生錯誤: {str(e)}'}), 500
        finally:
            db.close()

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

@app.route('/download-sales-detail/<filename>')
def download_sales_detail(filename):
    """下載生成的銷售明細表"""
    try:
        return send_from_directory(script_dir, filename, as_attachment=True)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 404

# --- Main Execution ---
if __name__ == '__main__':
    # This block is for local development testing only.
    # When deployed on Render with Gunicorn, this block will not be executed.
    # The scheduler thread is started below in the global scope.
    logging.info("Starting Flask development server for local testing...")
    app.run(host='0.0.0.0', port=5001, debug=False)

# --- Start the scheduler thread when the application module is loaded ---
# This ensures it runs even when started by Gunicorn on Render.
logging.info("Starting the background scheduler thread...")
scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
scheduler_thread.start()
logging.info("Background scheduler thread has been started.") 